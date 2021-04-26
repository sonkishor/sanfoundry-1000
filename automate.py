import requests
from bs4 import BeautifulSoup as bs
from pprint import pprint
import re
from nltk import flatten
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'sanf.xlsx'
ws1 = wb.active
url = "https://www.sanfoundry.com/1000-discrete-mathematics-questions-answers/"
# enter the URL of the subject you want to extract the MCQ of
request = requests.get(url)

soup = bs(request.content, 'html.parser')

table = soup.find_all('table', attrs = {'style':"width:100%"})
print(type(table))
urllist = []
for td in table:
    tr = td.find_all("li")
    for data in tr:
        datastr = str(data)
        link = re.findall(r"https://www.sanfoundry.com[a-z/-]*", datastr)
        urllist.append(link)
    flat = flatten(urllist)
y=2
z=2
column_question = 1
column_answer = 2
for hyp in flat:
    req = requests.get(hyp)
    reqSoup = bs(req.text, 'html.parser')
    
    #questions start here
    questions = []
    que = reqSoup.find("div", class_="entry-content", recursion=False).findNext('p').findNext('p')
    questions.append(que.contents[0])
    
    for i in range(9):
        q_next = que.findNext('p')
        questions.append(q_next.contents[0]) #final list
        que = q_next
        

    for k in range(len(questions)):
        ws1.cell(row=y, column=column_question).value = questions[k]
        y = y + 1
    y = y + 1
      
    #answers start here
    ans = reqSoup.find_all("div", class_ = "collapseomatic_content", recursive = True) #according to the lin
    i=0
    answers = []
    for a in ans:
        answers.append(a.text)
        i = i+1
        if i == 10:
            break
    print(len(questions))
    
    for l in range(len(answers)):
        ws1.cell(row=z, column=column_answer).value = answers[l]
        z = z + 1
    z = z + 1
wb.save("sanf.xlsx")       
